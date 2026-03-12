from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from .database import (
    PostgresManager, MSSQLManager,
    DatabaseError, DBConnectionError, QueryError, DBTimeoutError
)
from .config import Config
from datetime import datetime
from io import BytesIO
from concurrent.futures import ThreadPoolExecutor
import os
import math
import requests

app = Flask(__name__)

# Configure CORS with environment variable support
cors_origin = os.environ.get('CORS_ORIGIN', '*')
if cors_origin == '*':
    CORS(app)
else:
    CORS(app, origins=[cors_origin, 'http://localhost', 'http://127.0.0.1'])

# Initialize PostgreSQL manager
pg = PostgresManager(Config.DATABASE_URL)


def get_mssql_manager():
    """Get MS SQL manager from stored config."""
    config = pg.get_sql_config_with_password()
    if not config:
        return None
    return MSSQLManager(
        server=config['server'],
        database=config['database'],
        username=config['username'],
        password=config['password']
    )


def get_admin_db_name():
    """Get the admin database name from stored config."""
    config = pg.get_sql_config_with_password()
    return config.get('admin_database') if config else None


# Health check
@app.route('/health')
def health():
    return jsonify({'status': 'healthy', 'service': 'il-order-backend'})


# ============== Settings Endpoints ==============

@app.route('/api/settings', methods=['GET'])
def get_settings():
    """Get all application settings."""
    try:
        settings = pg.get_settings()
        return jsonify({
            'success': True,
            'settings': settings
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/settings', methods=['POST'])
def save_settings():
    """Save application settings."""
    try:
        data = request.get_json()
        pg.save_settings(data)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============== SQL Config Endpoints ==============

@app.route('/api/sql-config', methods=['GET'])
def get_sql_config():
    """Get SQL Server configuration (without password)."""
    try:
        config = pg.get_sql_config()
        return jsonify({
            'success': True,
            'config': dict(config) if config else None
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sql-config', methods=['POST'])
def save_sql_config():
    """Save SQL Server configuration."""
    try:
        data = request.get_json()
        required = ['server', 'database', 'username', 'password']
        for field in required:
            if not data.get(field):
                return jsonify({'success': False, 'error': f'Missing required field: {field}'}), 400

        config_id = pg.save_sql_config(
            server=data['server'],
            database=data['database'],
            username=data['username'],
            password=data['password'],
            name=data.get('name', 'default'),
            admin_database=data.get('admin_database')
        )

        return jsonify({'success': True, 'id': config_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/test-connection', methods=['POST'])
def test_connection():
    """Test MS SQL Server connection."""
    try:
        data = request.get_json()

        # Use provided credentials or stored config
        if data and all(k in data for k in ['server', 'database', 'username', 'password']):
            mssql = MSSQLManager(
                server=data['server'],
                database=data['database'],
                username=data['username'],
                password=data['password']
            )
        else:
            mssql = get_mssql_manager()
            if not mssql:
                return jsonify({
                    'success': False,
                    'error': 'No SQL configuration found. Please save configuration first.'
                }), 400

        result = mssql.test_connection()
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/test-admin-connection', methods=['POST'])
def test_admin_connection():
    """Test MS SQL Server admin database connection."""
    try:
        data = request.get_json()

        server = data.get('server')
        username = data.get('username')
        password = data.get('password')
        admin_database = data.get('admin_database')

        if not all([server, username, password, admin_database]):
            return jsonify({'success': False, 'error': 'Server, username, password, and admin database are required'}), 400

        mssql = MSSQLManager(
            server=server,
            database=admin_database,
            username=username,
            password=password
        )

        result = mssql.test_connection()
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============== Products Endpoints ==============

@app.route('/api/products', methods=['GET'])
def get_products():
    """Get products with optional search, filter, pagination, and sorting."""
    try:
        mssql = get_mssql_manager()
        if not mssql:
            return jsonify({'success': False, 'error': 'SQL Server not configured'}), 400

        search = request.args.get('search', '')
        limit = int(request.args.get('limit', 100))  # 0 = no limit (fetch all)
        offset = int(request.args.get('offset', 0))
        status_filter = request.args.get('filter', 'all')  # all, reorder, stocked (healthy also accepted)

        # Handle limit=0 as "no limit" for grouped view
        no_pagination = (limit == 0)
        show_excluded = request.args.get('show_excluded', 'false').lower() == 'true'
        sort_by = request.args.get('sort_by', 'description')  # upc, description, on_hand, threshold, monthly_avg, status
        sort_order = request.args.get('sort_order', 'asc')  # asc, desc

        # Get settings for threshold and order calculation
        settings = pg.get_settings()
        sales_period = int(settings.get('sales_period_days', 60))
        order_period_days = int(settings.get('order_period_days', 28))

        # Get overrides for threshold calculation
        overrides = {o['product_upc']: o for o in pg.get_all_product_overrides()}
        excluded_upcs = pg.get_excluded_upcs() if not show_excluded else set()

        # Get last supplier mapping for all products
        last_suppliers = mssql.get_last_suppliers_for_products()

        # Get quotations-in-progress quantities from admin DB
        admin_db_name = get_admin_db_name()
        qip_data = mssql.get_qip_quantities(admin_db_name) if admin_db_name else {}

        # For filtered views (reorder/healthy), we need to process all products
        # because threshold depends on PostgreSQL overrides
        if status_filter != 'all':
            # Use optimized combined query (single round-trip instead of two)
            result = mssql.get_products_with_sales(
                days=sales_period,
                search=search if search else None,
                sort_by=sort_by,
                sort_order=sort_order
            )
            all_products = result['products']

            # Enrich and filter products
            filtered = []
            for product in all_products:
                upc = product['ProductUPC']

                # Skip excluded products unless show_excluded is true
                is_excluded = upc in excluded_upcs
                if is_excluded and not show_excluded:
                    continue

                override = overrides.get(upc)
                monthly_avg = float(product.get('monthly_average') or 0)
                daily_avg = float(product.get('daily_average') or 0)
                dynamic_threshold = math.ceil(monthly_avg)

                # Determine effective threshold
                if override and override.get('manual_threshold') is not None:
                    threshold = override['manual_threshold']
                    threshold_type = 'manual'
                elif override and override.get('exclude_from_dynamic'):
                    threshold = product.get('ReorderLevel') or 0
                    threshold_type = 'system'
                else:
                    threshold = dynamic_threshold
                    threshold_type = 'dynamic'

                qty_on_hand = product.get('QuantOnHand') or 0
                pending_po_qty = product.get('pending_po_qty') or 0
                qip_qty = qip_data.get(upc, 0)
                effective_qty = qty_on_hand + pending_po_qty - qip_qty
                needs_reorder = effective_qty < threshold

                # Calculate suggested order quantity (same logic as needs-reorder endpoint)
                unit_qty2 = product.get('UnitQty2') or 1
                if unit_qty2 <= 0:
                    unit_qty2 = 1

                if needs_reorder:
                    effective_order_period = order_period_days
                    if override and override.get('manual_order_period_days'):
                        effective_order_period = override['manual_order_period_days']
                    projected_need = daily_avg * effective_order_period
                    cases_needed = int(-(-projected_need // unit_qty2))  # Ceiling division
                    suggested_qty = int(cases_needed * unit_qty2)
                    if override and override.get('manual_order_qty'):
                        suggested_qty = override['manual_order_qty']
                else:
                    suggested_qty = 0

                # Apply filter
                if status_filter == 'reorder' and not needs_reorder:
                    continue
                elif status_filter in ('stocked', 'healthy') and needs_reorder:
                    continue

                filtered.append({
                    'ProductID': product['ProductID'],
                    'ProductUPC': product['ProductUPC'],
                    'ProductSKU': product['ProductSKU'],
                    'ProductDescription': product['ProductDescription'],
                    'QuantOnHand': product['QuantOnHand'],
                    'QuantOnOrder': product['QuantOnOrder'],
                    'ReorderLevel': product['ReorderLevel'],
                    'ReorderQuant': product['ReorderQuant'],
                    'UnitCost': product['UnitCost'],
                    'UnitPrice': product['UnitPrice'],
                    'UnitQty2': product['UnitQty2'],
                    'UnitID2': product['UnitID2'],
                    'LastReceived': product['LastReceived'],
                    'LastSold': product['LastSold'],
                    'pending_po_qty': int(pending_po_qty),
                    'qip_qty': int(qip_qty),
                    'effective_qty': int(effective_qty),
                    'threshold': int(threshold),
                    'threshold_type': threshold_type,
                    'dynamic_threshold': int(dynamic_threshold),
                    'monthly_average': int(math.ceil(monthly_avg)),
                    'daily_average': round(daily_avg, 2),
                    'needs_reorder': needs_reorder,
                    'suggested_qty': suggested_qty,
                    'excluded': is_excluded or (override and override.get('exclude_from_orders', False)),
                    'override': override,
                    'last_supplier': last_suppliers.get(upc)
                })

            # Sort by fields that require Python-level sorting (not available in SQL)
            if sort_by == 'last_supplier':
                reverse_sort = sort_order == 'desc'
                filtered.sort(key=lambda x: (x.get('last_supplier') or '').lower(), reverse=reverse_sort)
            elif sort_by == 'status':
                # Sort by needs_reorder: True (Reorder) first when asc, False (OK) first when desc
                reverse_sort = sort_order == 'desc'
                filtered.sort(key=lambda x: (not x.get('needs_reorder', False)), reverse=reverse_sort)
            elif sort_by == 'threshold':
                # Sort by the calculated threshold value (not SQL ReorderLevel)
                reverse_sort = sort_order == 'desc'
                filtered.sort(key=lambda x: x.get('threshold', 0), reverse=reverse_sort)

            # Apply pagination to filtered results (unless no_pagination)
            total_count = len(filtered)
            if no_pagination:
                enriched = filtered
            else:
                enriched = filtered[offset:offset + limit]
        else:
            # "all" filter - use SQL-level pagination for maximum speed
            # Exception: when sorting by last_supplier, status, threshold, or no_pagination, we need to fetch all and sort in Python
            # (threshold requires Python sorting because actual threshold is calculated from overrides + dynamic values)
            use_python_pagination = sort_by in ('last_supplier', 'status', 'threshold') or no_pagination

            result = mssql.get_products_with_sales(
                days=sales_period,
                search=search if search else None,
                limit=None if use_python_pagination else limit,
                offset=None if use_python_pagination else offset,
                sort_by=sort_by if not use_python_pagination else 'description',
                sort_order=sort_order if not use_python_pagination else 'asc'
            )
            products = result['products']
            total_count = result['total_count']

            # Enrich products with threshold data
            all_enriched = []
            for product in products:
                upc = product['ProductUPC']

                # Skip excluded products unless show_excluded is true
                is_excluded = upc in excluded_upcs
                if is_excluded and not show_excluded:
                    total_count -= 1
                    continue

                override = overrides.get(upc)
                monthly_avg = float(product.get('monthly_average') or 0)
                daily_avg = float(product.get('daily_average') or 0)
                dynamic_threshold = math.ceil(monthly_avg)

                # Determine effective threshold
                if override and override.get('manual_threshold') is not None:
                    threshold = override['manual_threshold']
                    threshold_type = 'manual'
                elif override and override.get('exclude_from_dynamic'):
                    threshold = product.get('ReorderLevel') or 0
                    threshold_type = 'system'
                else:
                    threshold = dynamic_threshold
                    threshold_type = 'dynamic'

                qty_on_hand = product.get('QuantOnHand') or 0
                pending_po_qty = product.get('pending_po_qty') or 0
                qip_qty = qip_data.get(upc, 0)
                effective_qty = qty_on_hand + pending_po_qty - qip_qty
                needs_reorder = effective_qty < threshold

                # Calculate suggested order quantity (same logic as needs-reorder endpoint)
                unit_qty2 = product.get('UnitQty2') or 1
                if unit_qty2 <= 0:
                    unit_qty2 = 1

                if needs_reorder:
                    effective_order_period = order_period_days
                    if override and override.get('manual_order_period_days'):
                        effective_order_period = override['manual_order_period_days']
                    projected_need = daily_avg * effective_order_period
                    cases_needed = int(-(-projected_need // unit_qty2))  # Ceiling division
                    suggested_qty = int(cases_needed * unit_qty2)
                    if override and override.get('manual_order_qty'):
                        suggested_qty = override['manual_order_qty']
                else:
                    suggested_qty = 0

                all_enriched.append({
                    'ProductID': product['ProductID'],
                    'ProductUPC': product['ProductUPC'],
                    'ProductSKU': product['ProductSKU'],
                    'ProductDescription': product['ProductDescription'],
                    'QuantOnHand': product['QuantOnHand'],
                    'QuantOnOrder': product['QuantOnOrder'],
                    'ReorderLevel': product['ReorderLevel'],
                    'ReorderQuant': product['ReorderQuant'],
                    'UnitCost': product['UnitCost'],
                    'UnitPrice': product['UnitPrice'],
                    'UnitQty2': product['UnitQty2'],
                    'UnitID2': product['UnitID2'],
                    'LastReceived': product['LastReceived'],
                    'LastSold': product['LastSold'],
                    'pending_po_qty': int(pending_po_qty),
                    'qip_qty': int(qip_qty),
                    'effective_qty': int(effective_qty),
                    'threshold': int(threshold),
                    'threshold_type': threshold_type,
                    'dynamic_threshold': int(dynamic_threshold),
                    'monthly_average': int(math.ceil(monthly_avg)),
                    'daily_average': round(daily_avg, 2),
                    'needs_reorder': needs_reorder,
                    'suggested_qty': suggested_qty,
                    'excluded': is_excluded or (override and override.get('exclude_from_orders', False)),
                    'override': override,
                    'last_supplier': last_suppliers.get(upc)
                })

            # Sort by fields that require Python-level sorting, then apply pagination (unless no_pagination)
            if use_python_pagination:
                reverse_sort = sort_order == 'desc'
                if sort_by == 'last_supplier':
                    all_enriched.sort(key=lambda x: (x.get('last_supplier') or '').lower(), reverse=reverse_sort)
                elif sort_by == 'status':
                    # Sort by needs_reorder: True (Reorder) first when asc, False (OK) first when desc
                    all_enriched.sort(key=lambda x: (not x.get('needs_reorder', False)), reverse=reverse_sort)
                elif sort_by == 'threshold':
                    # Sort by the calculated threshold value (not SQL ReorderLevel)
                    all_enriched.sort(key=lambda x: x.get('threshold', 0), reverse=reverse_sort)
                total_count = len(all_enriched)
                if no_pagination:
                    enriched = all_enriched
                else:
                    enriched = all_enriched[offset:offset + limit]
            else:
                enriched = all_enriched

        return jsonify({
            'success': True,
            'products': enriched,
            'count': len(enriched),
            'total_count': total_count,
            'limit': limit,
            'offset': offset,
            'filter': status_filter
        })
    except DBTimeoutError as e:
        return jsonify({'success': False, 'error': f'Database timeout: {e}'}), 504
    except DBConnectionError as e:
        return jsonify({'success': False, 'error': f'Connection error: {e}'}), 503
    except QueryError as e:
        return jsonify({'success': False, 'error': f'Query error: {e}'}), 500
    except DatabaseError as e:
        return jsonify({'success': False, 'error': f'Database error: {e}'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/products/<upc>', methods=['GET'])
def get_product(upc):
    """Get single product details."""
    try:
        mssql = get_mssql_manager()
        if not mssql:
            return jsonify({'success': False, 'error': 'SQL Server not configured'}), 400

        product = mssql.get_product_by_upc(upc)
        if not product:
            return jsonify({'success': False, 'error': 'Product not found'}), 404

        settings = pg.get_settings()
        sales_period = int(settings.get('sales_period_days', 60))
        sales_data = mssql.get_sales_data(upc, sales_period)
        override = pg.get_product_override(upc)

        # Get QIP quantity for this product
        admin_db_name = get_admin_db_name()
        qip_data = mssql.get_qip_quantities(admin_db_name) if admin_db_name else {}
        qip_qty = qip_data.get(upc, 0)

        # Adjust effective_qty with QIP
        product = dict(product)
        qty_on_hand = product.get('QuantOnHand') or 0
        pending_po_qty = product.get('pending_po_qty') or 0
        product['qip_qty'] = qip_qty
        product['effective_qty'] = qty_on_hand + pending_po_qty - qip_qty

        return jsonify({
            'success': True,
            'product': product,
            'sales_data': sales_data,
            'override': override
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/products/<upc>/override', methods=['POST'])
def save_product_override(upc):
    """Save product override settings."""
    try:
        data = request.get_json()

        override_id = pg.save_product_override(
            product_upc=upc,
            exclude_from_dynamic=data.get('exclude_from_dynamic', False),
            exclude_from_orders=data.get('exclude_from_orders', False),
            manual_threshold=data.get('manual_threshold'),
            manual_order_qty=data.get('manual_order_qty'),
            manual_order_period_days=data.get('manual_order_period_days'),
            manual_unit_cost=data.get('manual_unit_cost'),
            notes=data.get('notes')
        )

        return jsonify({'success': True, 'id': override_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/products/<upc>/override', methods=['DELETE'])
def delete_product_override(upc):
    """Delete product override."""
    try:
        deleted = pg.delete_product_override(upc)
        return jsonify({'success': True, 'deleted': deleted})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/products/<upc>/unit-qty2', methods=['PUT'])
def update_unit_qty2(upc):
    """Update UnitQty2 for a product."""
    try:
        mssql = get_mssql_manager()
        if not mssql:
            return jsonify({'success': False, 'error': 'SQL Server not configured'}), 400

        data = request.get_json()
        unit_qty2 = data.get('unit_qty2')

        if unit_qty2 is None or unit_qty2 < 1:
            return jsonify({'success': False, 'error': 'Invalid unit_qty2 value'}), 400

        updated = mssql.update_unit_qty2(upc, int(unit_qty2))
        return jsonify({'success': True, 'updated': updated})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/products/<upc>/exclude', methods=['POST'])
def toggle_product_exclusion(upc):
    """Toggle product exclusion from orders."""
    try:
        new_state = pg.toggle_product_exclusion(upc)
        return jsonify({
            'success': True,
            'excluded': new_state,
            'upc': upc
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/products/excluded', methods=['GET'])
def get_excluded_products():
    """Get list of products excluded from orders with descriptions."""
    try:
        excluded = pg.get_excluded_products()
        excluded_list = [dict(p) for p in excluded]

        mssql = get_mssql_manager()
        sql_configured = mssql is not None

        if sql_configured and excluded_list:
            for item in excluded_list:
                product = mssql.get_product_by_upc(item['product_upc'])
                if product:
                    item['description'] = product.get('ProductDescription', '')
                else:
                    item['description'] = '(Product not found)'

        return jsonify({
            'success': True,
            'products': excluded_list,
            'count': len(excluded_list),
            'sql_configured': sql_configured
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============== Suppliers Endpoints ==============

@app.route('/api/suppliers', methods=['GET'])
def get_suppliers():
    """Get suppliers from purchase history."""
    try:
        mssql = get_mssql_manager()
        if not mssql:
            return jsonify({'success': False, 'error': 'SQL Server not configured'}), 400

        suppliers = mssql.get_suppliers_from_purchase_history()
        return jsonify({
            'success': True,
            'suppliers': suppliers
        })
    except DBTimeoutError as e:
        return jsonify({'success': False, 'error': f'Database timeout: {e}'}), 504
    except DBConnectionError as e:
        return jsonify({'success': False, 'error': f'Connection error: {e}'}), 503
    except QueryError as e:
        return jsonify({'success': False, 'error': f'Query error: {e}'}), 500
    except DatabaseError as e:
        return jsonify({'success': False, 'error': f'Database error: {e}'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/suppliers/excluded', methods=['GET'])
def get_excluded_suppliers():
    """Get list of excluded suppliers from grouped view."""
    try:
        excluded = pg.get_excluded_suppliers()
        return jsonify({
            'success': True,
            'suppliers': [dict(s) for s in excluded],
            'count': len(excluded)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/suppliers/exclude', methods=['POST'])
def exclude_supplier():
    """Add a supplier to the exclusion list."""
    try:
        data = request.get_json()
        supplier_name = data.get('supplier_name')

        if not supplier_name:
            return jsonify({'success': False, 'error': 'supplier_name is required'}), 400

        result_id = pg.exclude_supplier(supplier_name)
        return jsonify({
            'success': True,
            'id': result_id,
            'supplier_name': supplier_name
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/suppliers/include', methods=['POST'])
def include_supplier():
    """Remove a supplier from the exclusion list."""
    try:
        data = request.get_json()
        supplier_name = data.get('supplier_name')

        if not supplier_name:
            return jsonify({'success': False, 'error': 'supplier_name is required'}), 400

        deleted = pg.include_supplier(supplier_name)
        return jsonify({
            'success': True,
            'deleted': deleted,
            'supplier_name': supplier_name
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/products/by-supplier/<int:supplier_id>', methods=['GET'])
def get_products_by_supplier(supplier_id):
    """Get products ordered from a specific supplier."""
    try:
        mssql = get_mssql_manager()
        if not mssql:
            return jsonify({'success': False, 'error': 'SQL Server not configured'}), 400

        products = mssql.get_products_by_supplier(supplier_id)

        # Filter out excluded products
        excluded_upcs = pg.get_excluded_upcs()
        products = [p for p in products if p['ProductUPC'] not in excluded_upcs]

        return jsonify({
            'success': True,
            'products': products,
            'count': len(products)
        })
    except DBTimeoutError as e:
        return jsonify({'success': False, 'error': f'Database timeout: {e}'}), 504
    except DBConnectionError as e:
        return jsonify({'success': False, 'error': f'Connection error: {e}'}), 503
    except QueryError as e:
        return jsonify({'success': False, 'error': f'Query error: {e}'}), 500
    except DatabaseError as e:
        return jsonify({'success': False, 'error': f'Database error: {e}'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/suppliers/<int:supplier_id>/details', methods=['GET'])
def get_supplier_details(supplier_id):
    """Get supplier details for PO ship-to fields."""
    try:
        mssql = get_mssql_manager()
        if not mssql:
            return jsonify({'success': False, 'error': 'SQL Server not configured'}), 400

        supplier = mssql.get_supplier_details(supplier_id)
        if not supplier:
            return jsonify({'success': False, 'error': 'Supplier not found'}), 404

        return jsonify({
            'success': True,
            'supplier': dict(supplier)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============== Purchase Order Endpoints ==============

@app.route('/api/po/validate-number', methods=['POST'])
def validate_po_number():
    """Validate if a PO number is available (not already used)."""
    try:
        mssql = get_mssql_manager()
        if not mssql:
            return jsonify({'success': False, 'error': 'SQL Server not configured'}), 400

        data = request.get_json()
        po_number = data.get('po_number', '').strip()

        if not po_number:
            return jsonify({'success': True, 'valid': False, 'message': 'PO number is required'})

        if len(po_number) > 20:
            return jsonify({'success': True, 'valid': False, 'message': 'PO number must be 20 characters or less'})

        result = mssql.validate_po_number(po_number)
        return jsonify({
            'success': True,
            'valid': result['valid'],
            'message': result['message']
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/orders/<int:order_id>/create-po', methods=['POST'])
def create_purchase_order(order_id):
    """Create a Purchase Order in MS SQL from an order draft."""
    try:
        mssql = get_mssql_manager()
        if not mssql:
            return jsonify({'success': False, 'error': 'SQL Server not configured'}), 400

        # Get the order draft
        order = pg.get_order_draft(order_id)
        if not order:
            return jsonify({'success': False, 'error': 'Order not found'}), 404

        data = request.get_json()
        po_number = data.get('po_number', '').strip()
        supplier_id = data.get('supplier_id') or order.get('supplier_id')

        # Validate PO number
        if not po_number:
            return jsonify({'success': False, 'error': 'PO number is required'}), 400

        if len(po_number) > 20:
            return jsonify({'success': False, 'error': 'PO number must be 20 characters or less'}), 400

        validation = mssql.validate_po_number(po_number)
        if not validation['valid']:
            return jsonify({'success': False, 'error': validation['message']}), 400

        # Validate supplier
        if not supplier_id:
            return jsonify({'success': False, 'error': 'Supplier is required'}), 400

        # Get supplier details for ship-to fields
        supplier = mssql.get_supplier_details(int(supplier_id))
        if not supplier:
            return jsonify({'success': False, 'error': 'Supplier not found'}), 404

        # Get order items
        items = pg.get_order_draft_items(order_id)
        if not items:
            return jsonify({'success': False, 'error': 'Order has no items'}), 400

        # Get product details from MS SQL for PO line items
        upcs = [item['product_upc'] for item in items]
        products = mssql.get_products_for_po(upcs)

        # Check for missing products
        missing_upcs = [upc for upc in upcs if upc not in products]
        if missing_upcs:
            return jsonify({
                'success': False,
                'error': f'Products not found in system: {", ".join(missing_upcs[:5])}{"..." if len(missing_upcs) > 5 else ""}'
            }), 400

        # Prepare PO header data
        po_data = {
            'po_number': po_number,
            'supplier_id': int(supplier_id),
            'business_name': supplier.get('BusinessName', ''),
            'account_no': supplier.get('AccountNo', ''),
            'po_title': order.get('name') or 'IL-Order Export',
            'shipto': supplier.get('BusinessName', ''),
            'ship_address1': supplier.get('Address1', ''),
            'ship_address2': supplier.get('Address2', ''),
            'ship_contact': supplier.get('Contactname', ''),
            'ship_city': supplier.get('City', ''),
            'ship_state': supplier.get('State', ''),
            'ship_zipcode': supplier.get('ZipCode', ''),
            'ship_phone': supplier.get('Phone_Number', '')
        }

        # Prepare PO line items
        po_items = []
        for item in items:
            upc = item['product_upc']
            product = products.get(upc, {})
            qty_ordered = item['final_qty'] or 0
            unit_cost = float(item['unit_cost'] or 0)

            po_items.append({
                'product_id': product.get('ProductID'),
                'cate_id': product.get('CateID'),
                'sub_cate_id': product.get('SubCateID'),
                'unit_desc': product.get('UnitDesc', ''),
                'product_sku': product.get('ProductSKU', ''),
                'product_upc': upc,
                'product_description': item['product_description'] or '',
                'item_weight': product.get('ItemWeight'),
                'qty_ordered': qty_ordered,
                'unit_cost': unit_cost,
                'extended_cost': qty_ordered * unit_cost
            })

        # Create the purchase order
        result = mssql.create_purchase_order(po_data, po_items)

        if result.get('success'):
            return jsonify({
                'success': True,
                'po_id': result['po_id'],
                'po_number': result['po_number']
            })
        else:
            return jsonify({'success': False, 'error': 'Failed to create purchase order'}), 500

    except QueryError as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============== Analysis Endpoints ==============

@app.route('/api/analysis/needs-reorder', methods=['GET'])
def get_needs_reorder():
    """Get products with reorder analysis. Supports filtering by status."""
    try:
        mssql = get_mssql_manager()
        if not mssql:
            return jsonify({'success': False, 'error': 'SQL Server not configured'}), 400

        supplier_id = request.args.get('supplier_id', type=int)
        filter_mode = request.args.get('filter', 'all')
        sort_by = request.args.get('sort_by')  # None = use default sort; options: status, upc, description, on_hand, threshold, suggested_qty, cases
        sort_order = request.args.get('sort_order', 'asc')  # asc, desc

        if supplier_id:
            products = mssql.get_products_by_supplier(supplier_id)
        else:
            products = mssql.get_all_products()

        settings = pg.get_settings()
        sales_period = int(settings.get('sales_period_days', 60))
        order_period_days = int(settings.get('order_period_days', 28))

        result_products = []
        overrides = {o['product_upc']: o for o in pg.get_all_product_overrides()}
        excluded_upcs = pg.get_excluded_upcs()

        all_sales_data = mssql.get_all_sales_data(sales_period)
        last_suppliers = mssql.get_last_suppliers_for_products()

        # Get quotations-in-progress quantities from admin DB
        admin_db_name = get_admin_db_name()
        qip_data = mssql.get_qip_quantities(admin_db_name) if admin_db_name else {}

        for product in products:
            upc = product['ProductUPC']

            if upc in excluded_upcs:
                continue

            override = overrides.get(upc)

            sales_data = all_sales_data.get(upc, {'monthly_average': 0, 'daily_average': 0})
            dynamic_threshold = math.ceil(sales_data['monthly_average'])

            if override and override.get('manual_threshold') is not None:
                threshold = override['manual_threshold']
            elif override and override.get('exclude_from_dynamic'):
                threshold = product.get('ReorderLevel') or 0
            else:
                threshold = dynamic_threshold

            qty_on_hand = product.get('QuantOnHand') or 0
            pending_po_qty = product.get('pending_po_qty') or 0
            qip_qty = qip_data.get(upc, 0)
            effective_qty = qty_on_hand + pending_po_qty - qip_qty
            needs_reorder = effective_qty < threshold

            unit_qty2 = product.get('UnitQty2') or 1
            if unit_qty2 <= 0:
                unit_qty2 = 1

            if needs_reorder:
                daily_avg = sales_data['daily_average']
                effective_order_period = order_period_days
                if override and override.get('manual_order_period_days'):
                    effective_order_period = override['manual_order_period_days']
                projected_need = daily_avg * effective_order_period
                cases_needed = -(-projected_need // unit_qty2)
                suggested_qty = int(cases_needed * unit_qty2)

                if override and override.get('manual_order_qty'):
                    suggested_qty = override['manual_order_qty']
                    cases_needed = -(-suggested_qty // unit_qty2)
            else:
                suggested_qty = 0
                cases_needed = 0

            # Determine effective unit cost (override or system)
            effective_unit_cost = product.get('UnitCost') or 0
            if override and override.get('manual_unit_cost') is not None:
                effective_unit_cost = float(override['manual_unit_cost'])

            product_data = {
                **product,
                'pending_po_qty': int(pending_po_qty),
                'qip_qty': int(qip_qty),
                'effective_qty': int(effective_qty),
                'threshold': int(threshold),
                'monthly_average': int(math.ceil(sales_data['monthly_average'])),
                'daily_average': round(sales_data['daily_average'], 2),
                'suggested_qty': suggested_qty,
                'cases_needed': int(cases_needed),
                'unit_qty2': unit_qty2,
                'deficit': int(threshold - effective_qty),
                'status': 'needs_reorder' if needs_reorder else 'sufficient',
                'effective_unit_cost': effective_unit_cost,
                'has_cost_override': override and override.get('manual_unit_cost') is not None,
                'last_supplier': last_suppliers.get(upc)
            }

            if filter_mode == 'all':
                result_products.append(product_data)
            elif filter_mode == 'needs_reorder' and needs_reorder:
                result_products.append(product_data)
            elif filter_mode == 'sufficient' and not needs_reorder:
                result_products.append(product_data)

        # Sort products
        if sort_by:
            def get_sort_key(p):
                key_map = {
                    'status': 0 if p['status'] == 'needs_reorder' else 1,
                    'upc': (p.get('ProductUPC') or '').lower(),
                    'description': (p.get('ProductDescription') or '').lower(),
                    'on_hand': p.get('QuantOnHand') or 0,
                    'case_qty': p.get('unit_qty2') or 0,
                    'threshold': p.get('threshold') or 0,
                    'suggested_qty': p.get('suggested_qty') or 0,
                    'cases': p.get('cases_needed') or 0
                }
                return key_map.get(sort_by, (p.get('ProductDescription') or '').lower())

            reverse_sort = sort_order == 'desc'
            result_products.sort(key=get_sort_key, reverse=reverse_sort)
        else:
            # Default: needs_reorder first, then by deficit, then by description
            result_products.sort(key=lambda x: (
                0 if x['status'] == 'needs_reorder' else 1,
                -x['deficit'] if x['status'] == 'needs_reorder' else 0,
                (x.get('ProductDescription') or '').lower()
            ))

        return jsonify({
            'success': True,
            'products': result_products,
            'count': len(result_products),
            'filter': filter_mode
        })
    except DBTimeoutError as e:
        return jsonify({'success': False, 'error': f'Database timeout: {e}'}), 504
    except DBConnectionError as e:
        return jsonify({'success': False, 'error': f'Connection error: {e}'}), 503
    except QueryError as e:
        return jsonify({'success': False, 'error': f'Query error: {e}'}), 500
    except DatabaseError as e:
        return jsonify({'success': False, 'error': f'Database error: {e}'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/analysis/summary', methods=['GET'])
def get_summary():
    """Get dashboard summary statistics."""
    try:
        mssql = get_mssql_manager()
        if not mssql:
            return jsonify({
                'success': True,
                'summary': {
                    'configured': False
                }
            })

        products = mssql.get_all_products()
        settings = pg.get_settings()
        sales_period = int(settings.get('sales_period_days', 60))

        overrides = {o['product_upc']: o for o in pg.get_all_product_overrides()}
        excluded_upcs = pg.get_excluded_upcs()

        # Filter out excluded products
        active_products = [p for p in products if p['ProductUPC'] not in excluded_upcs]

        total_products = len(active_products)
        needs_reorder_count = 0

        # Get all sales data in a single query (optimized)
        all_sales_data = mssql.get_all_sales_data(sales_period)

        # Get quotations-in-progress quantities from admin DB
        admin_db_name = get_admin_db_name()
        qip_data = mssql.get_qip_quantities(admin_db_name) if admin_db_name else {}

        for product in active_products:
            upc = product['ProductUPC']
            override = overrides.get(upc)

            # Look up sales data from batch result
            sales_data = all_sales_data.get(upc, {'monthly_average': 0})
            dynamic_threshold = math.ceil(sales_data['monthly_average'])

            if override and override.get('manual_threshold') is not None:
                threshold = override['manual_threshold']
            elif override and override.get('exclude_from_dynamic'):
                threshold = product.get('ReorderLevel') or 0
            else:
                threshold = dynamic_threshold

            qty_on_hand = product.get('QuantOnHand') or 0
            pending_po_qty = product.get('pending_po_qty') or 0
            qip_qty = qip_data.get(upc, 0)
            effective_qty = qty_on_hand + pending_po_qty - qip_qty

            if effective_qty < threshold:
                needs_reorder_count += 1

        return jsonify({
            'success': True,
            'summary': {
                'configured': True,
                'total_products': total_products,
                'needs_reorder': needs_reorder_count,
                'healthy': total_products - needs_reorder_count,
                'settings': settings
            }
        })
    except DBTimeoutError as e:
        return jsonify({'success': False, 'error': f'Database timeout: {e}'}), 504
    except DBConnectionError as e:
        return jsonify({'success': False, 'error': f'Connection error: {e}'}), 503
    except QueryError as e:
        return jsonify({'success': False, 'error': f'Query error: {e}'}), 500
    except DatabaseError as e:
        return jsonify({'success': False, 'error': f'Database error: {e}'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============== Order Draft Endpoints ==============

@app.route('/api/orders', methods=['GET'])
def get_orders():
    """Get all order drafts."""
    try:
        status = request.args.get('status')
        orders = pg.get_order_drafts(status)
        return jsonify({
            'success': True,
            'orders': [dict(o) for o in orders]
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/orders', methods=['POST'])
def create_order():
    """Create a new order draft with items."""
    try:
        data = request.get_json()
        name = data.get('name', f"Order {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        supplier_id = data.get('supplier_id')
        supplier_name = data.get('supplier_name')
        items = data.get('items', [])

        order_id = pg.create_order_draft(name=name, supplier_id=supplier_id, supplier_name=supplier_name)

        for item in items:
            pg.add_order_draft_item(
                order_id=order_id,
                product_upc=item['upc'],
                product_description=item.get('description', ''),
                current_qty=item.get('on_hand', 0),
                threshold=item.get('threshold', 0),
                suggested_qty=item.get('suggested_qty', 0),
                final_qty=item.get('order_qty', item.get('suggested_qty', 0)),
                unit_qty2=item.get('unit_qty2', 1),
                unit_cost=item.get('unit_cost', 0)
            )

        return jsonify({'success': True, 'id': order_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/orders/<int:order_id>', methods=['GET'])
def get_order(order_id):
    """Get order draft details with items."""
    try:
        order = pg.get_order_draft(order_id)
        if not order:
            return jsonify({'success': False, 'error': 'Order not found'}), 404

        items = pg.get_order_draft_items(order_id)

        total_items = len(items)
        total_qty = sum(i['final_qty'] or 0 for i in items)
        total_cost = sum((i['final_qty'] or 0) * float(i['unit_cost'] or 0) for i in items)

        return jsonify({
            'success': True,
            'order': dict(order),
            'items': [dict(i) for i in items],
            'summary': {
                'total_items': total_items,
                'total_qty': total_qty,
                'total_cost': round(total_cost, 2)
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/orders/<int:order_id>', methods=['PUT'])
def update_order(order_id):
    """Update order draft."""
    try:
        data = request.get_json()

        if 'name' in data:
            pg.update_order_draft(order_id, name=data['name'])

        if 'status' in data:
            pg.update_order_draft_status(order_id, data['status'])

        if 'items' in data:
            for item in data['items']:
                if 'id' in item and 'final_qty' in item:
                    pg.update_order_draft_item(item['id'], item['final_qty'])

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/orders/<int:order_id>', methods=['DELETE'])
def delete_order(order_id):
    """Delete order draft."""
    try:
        deleted = pg.delete_order_draft(order_id)
        return jsonify({'success': True, 'deleted': deleted})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/orders/<int:order_id>/items/<int:item_id>', methods=['DELETE'])
def delete_order_item(order_id, item_id):
    """Delete an item from an order draft."""
    try:
        deleted = pg.delete_order_draft_item(item_id)
        return jsonify({'success': True, 'deleted': deleted})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============== Export Endpoints ==============

@app.route('/api/orders/<int:order_id>/export/excel', methods=['GET'])
def export_order_excel(order_id):
    """Export order to Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        order = pg.get_order_draft(order_id)
        if not order:
            return jsonify({'success': False, 'error': 'Order not found'}), 404

        items = pg.get_order_draft_items(order_id)

        # Parse columns parameter
        columns_param = request.args.get('columns', '')
        all_columns = ['upc', 'description', 'on_hand', 'threshold', 'suggested_qty', 'order_qty', 'cases', 'unit_cost', 'total']
        if columns_param:
            selected_columns = [c.strip() for c in columns_param.split(',') if c.strip() in all_columns]
        else:
            selected_columns = all_columns

        # Column configuration
        column_config = {
            'upc': {'header': 'UPC', 'width': 15},
            'description': {'header': 'Description', 'width': 40},
            'on_hand': {'header': 'On Hand', 'width': 12},
            'threshold': {'header': 'Threshold', 'width': 12},
            'suggested_qty': {'header': 'Suggested Qty', 'width': 14},
            'order_qty': {'header': 'Order Qty', 'width': 12},
            'cases': {'header': 'Cases', 'width': 10},
            'unit_cost': {'header': 'Unit Cost', 'width': 12},
            'total': {'header': 'Total', 'width': 12},
        }

        wb = Workbook()
        ws = wb.active
        ws.title = "Order"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws['A1'] = f"Order: {order['name'] or 'Untitled'}"
        ws['A1'].font = Font(bold=True, size=14)

        # Write headers for selected columns only
        headers = [column_config[col]['header'] for col in selected_columns]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Write data rows
        for row_idx, item in enumerate(items, 4):
            unit_qty2 = float(item['unit_qty2'] or 1)
            cases = int((item['final_qty'] or 0) / unit_qty2) if unit_qty2 > 0 else 0
            total = (item['final_qty'] or 0) * float(item['unit_cost'] or 0)

            # Build row data based on selected columns
            row_data = {
                'upc': item['product_upc'],
                'description': item['product_description'],
                'on_hand': item['current_qty'],
                'threshold': item['threshold'],
                'suggested_qty': item['suggested_qty'],
                'order_qty': item['final_qty'],
                'cases': cases,
                'unit_cost': float(item['unit_cost'] or 0),
                'total': total,
            }

            for col_idx, col_id in enumerate(selected_columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data[col_id])
                cell.border = thin_border
                if col_id in ['unit_cost', 'total']:
                    cell.number_format = '$#,##0.00'

        # Summary row - only show if relevant columns are selected
        summary_row = len(items) + 5
        if 'order_qty' in selected_columns:
            order_qty_idx = selected_columns.index('order_qty') + 1
            # Put TOTALS label in column before order_qty, or in order_qty column if it's first
            if order_qty_idx > 1:
                ws.cell(row=summary_row, column=order_qty_idx - 1, value="TOTALS:").font = Font(bold=True)
            ws.cell(row=summary_row, column=order_qty_idx, value=sum(i['final_qty'] or 0 for i in items)).font = Font(bold=True)

        if 'total' in selected_columns:
            total_idx = selected_columns.index('total') + 1
            total_cost = sum((i['final_qty'] or 0) * float(i['unit_cost'] or 0) for i in items)
            ws.cell(row=summary_row, column=total_idx, value=total_cost).font = Font(bold=True)
            ws.cell(row=summary_row, column=total_idx).number_format = '$#,##0.00'

        # Set column widths for selected columns
        col_letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        for col_idx, col_id in enumerate(selected_columns):
            ws.column_dimensions[col_letters[col_idx]].width = column_config[col_id]['width']

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        pg.update_order_draft_status(order_id, 'exported')

        filename = f"order_{order_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except ImportError:
        return jsonify({'success': False, 'error': 'openpyxl not installed'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/orders/<int:order_id>/export/pdf', methods=['GET'])
def export_order_pdf(order_id):
    """Export order to PDF file."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        order = pg.get_order_draft(order_id)
        if not order:
            return jsonify({'success': False, 'error': 'Order not found'}), 404

        items = pg.get_order_draft_items(order_id)

        # Parse columns parameter
        columns_param = request.args.get('columns', '')
        all_columns = ['upc', 'description', 'on_hand', 'threshold', 'suggested_qty', 'order_qty', 'cases', 'unit_cost', 'total']
        if columns_param:
            selected_columns = [c.strip() for c in columns_param.split(',') if c.strip() in all_columns]
        else:
            selected_columns = all_columns

        # Column configuration for PDF
        column_config = {
            'upc': {'header': 'UPC', 'width': 1.1*inch},
            'description': {'header': 'Description', 'width': 2.5*inch},
            'on_hand': {'header': 'On Hand', 'width': 0.7*inch},
            'threshold': {'header': 'Threshold', 'width': 0.8*inch},
            'suggested_qty': {'header': 'Suggested', 'width': 0.8*inch},
            'order_qty': {'header': 'Order Qty', 'width': 0.8*inch},
            'cases': {'header': 'Cases', 'width': 0.6*inch},
            'unit_cost': {'header': 'Unit Cost', 'width': 0.8*inch},
            'total': {'header': 'Total', 'width': 0.9*inch},
        }

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, spaceAfter=12)
        elements.append(Paragraph(f"Order: {order['name'] or 'Untitled'}", title_style))
        elements.append(Spacer(1, 0.2*inch))

        # Build headers for selected columns
        headers = [column_config[col]['header'] for col in selected_columns]
        data = [headers]

        for item in items:
            unit_qty2 = float(item['unit_qty2'] or 1)
            cases = int((item['final_qty'] or 0) / unit_qty2) if unit_qty2 > 0 else 0
            total = (item['final_qty'] or 0) * float(item['unit_cost'] or 0)

            # Build row data based on selected columns
            row_data = {
                'upc': item['product_upc'] or '',
                'description': (item['product_description'] or '')[:35],
                'on_hand': str(int(item['current_qty'] or 0)),
                'threshold': str(int(item['threshold'] or 0)),
                'suggested_qty': str(int(item['suggested_qty'] or 0)),
                'order_qty': str(int(item['final_qty'] or 0)),
                'cases': str(cases),
                'unit_cost': f"${float(item['unit_cost'] or 0):.2f}",
                'total': f"${total:.2f}",
            }

            data.append([row_data[col] for col in selected_columns])

        # Build totals row
        total_qty = sum(i['final_qty'] or 0 for i in items)
        total_cost = sum((i['final_qty'] or 0) * float(i['unit_cost'] or 0) for i in items)

        totals_row = []
        for col in selected_columns:
            if col == 'order_qty':
                totals_row.append(str(int(total_qty)))
            elif col == 'total':
                totals_row.append(f"${total_cost:.2f}")
            elif col == 'suggested_qty':
                totals_row.append('TOTALS:')
            else:
                totals_row.append('')
        data.append(totals_row)

        # Get column widths for selected columns
        col_widths = [column_config[col]['width'] for col in selected_columns]

        table = Table(data, colWidths=col_widths)

        # Find description column index for left-alignment
        desc_col_idx = selected_columns.index('description') if 'description' in selected_columns else -1

        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a73e8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
        ]

        # Left-align description column if present
        if desc_col_idx >= 0:
            table_style.append(('ALIGN', (desc_col_idx, 1), (desc_col_idx, -1), 'LEFT'))

        table.setStyle(TableStyle(table_style))

        elements.append(table)
        doc.build(elements)
        output.seek(0)

        pg.update_order_draft_status(order_id, 'exported')

        filename = f"order_{order_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    except ImportError:
        return jsonify({'success': False, 'error': 'reportlab not installed'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============== Inventory Direct Export Endpoints ==============

@app.route('/api/inventory/export/excel', methods=['POST'])
def export_inventory_excel():
    """Export selected inventory items directly to Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        data = request.get_json()
        items = data.get('items', [])
        columns_list = data.get('columns')

        if not items:
            return jsonify({'success': False, 'error': 'No items provided'}), 400

        # Column configuration
        all_columns = ['upc', 'description', 'on_hand', 'threshold', 'suggested_qty', 'order_qty', 'cases', 'unit_cost', 'total']
        column_config = {
            'upc': {'header': 'UPC', 'width': 15},
            'description': {'header': 'Description', 'width': 40},
            'on_hand': {'header': 'On Hand', 'width': 12},
            'threshold': {'header': 'Threshold', 'width': 12},
            'suggested_qty': {'header': 'Suggested Qty', 'width': 14},
            'order_qty': {'header': 'Order Qty', 'width': 12},
            'cases': {'header': 'Cases', 'width': 10},
            'unit_cost': {'header': 'Unit Cost', 'width': 12},
            'total': {'header': 'Total', 'width': 12},
        }

        if columns_list:
            selected_columns = [c for c in columns_list if c in all_columns]
        else:
            selected_columns = all_columns

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Export"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a73e8", end_color="1a73e8", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Calculate totals for footer
        total_units = sum(item.get('order_qty', 0) for item in items)
        total_cost = sum(item.get('order_qty', 0) * float(item.get('unit_cost', 0)) for item in items)

        # Headers - start at row 1
        headers = [column_config[col]['header'] for col in selected_columns]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Data rows - start at row 2
        for row_idx, item in enumerate(items, 2):
            order_qty = item.get('order_qty', 0)
            unit_cost = float(item.get('unit_cost', 0))
            cases = item.get('cases', 0)
            total = order_qty * unit_cost

            row_data = {
                'upc': item.get('upc', ''),
                'description': item.get('description', ''),
                'on_hand': item.get('on_hand', 0),
                'threshold': item.get('threshold', 0),
                'suggested_qty': item.get('suggested_qty', 0),
                'order_qty': order_qty,
                'cases': cases,
                'unit_cost': unit_cost,
                'total': total
            }

            for col_idx, col_id in enumerate(selected_columns, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data[col_id])
                cell.border = thin_border
                if col_id in ['unit_cost', 'total']:
                    cell.number_format = '$#,##0.00'

        # Summary row - only show if relevant columns are selected
        summary_row = len(items) + 2
        if 'order_qty' in selected_columns:
            order_qty_idx = selected_columns.index('order_qty') + 1
            if order_qty_idx > 1:
                ws.cell(row=summary_row, column=order_qty_idx - 1, value="TOTALS:").font = Font(bold=True)
            ws.cell(row=summary_row, column=order_qty_idx, value=total_units).font = Font(bold=True)

        if 'total' in selected_columns:
            total_idx = selected_columns.index('total') + 1
            ws.cell(row=summary_row, column=total_idx, value=total_cost).font = Font(bold=True)
            ws.cell(row=summary_row, column=total_idx).number_format = '$#,##0.00'

        # Column widths for selected columns
        col_letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        for col_idx, col_id in enumerate(selected_columns):
            ws.column_dimensions[col_letters[col_idx]].width = column_config[col_id]['width']

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"inventory-export-{datetime.now().strftime('%Y%m%d')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except ImportError:
        return jsonify({'success': False, 'error': 'openpyxl not installed'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/inventory/export/pdf', methods=['POST'])
def export_inventory_pdf():
    """Export selected inventory items directly to PDF file."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        data = request.get_json()
        items = data.get('items', [])
        columns_list = data.get('columns')

        if not items:
            return jsonify({'success': False, 'error': 'No items provided'}), 400

        # Column configuration for PDF
        all_columns = ['upc', 'description', 'on_hand', 'threshold', 'suggested_qty', 'order_qty', 'cases', 'unit_cost', 'total']
        column_config = {
            'upc': {'header': 'UPC', 'width': 1.1*inch},
            'description': {'header': 'Description', 'width': 2.5*inch},
            'on_hand': {'header': 'On Hand', 'width': 0.7*inch},
            'threshold': {'header': 'Threshold', 'width': 0.8*inch},
            'suggested_qty': {'header': 'Suggested', 'width': 0.8*inch},
            'order_qty': {'header': 'Order Qty', 'width': 0.8*inch},
            'cases': {'header': 'Cases', 'width': 0.6*inch},
            'unit_cost': {'header': 'Unit Cost', 'width': 0.8*inch},
            'total': {'header': 'Total', 'width': 0.9*inch},
        }

        if columns_list:
            selected_columns = [c for c in columns_list if c in all_columns]
        else:
            selected_columns = all_columns

        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []

        # Calculate totals for footer
        total_units = sum(item.get('order_qty', 0) for item in items)
        total_cost = sum(item.get('order_qty', 0) * float(item.get('unit_cost', 0)) for item in items)

        # Build headers for selected columns
        headers = [column_config[col]['header'] for col in selected_columns]
        table_data = [headers]

        # Table rows
        for item in items:
            order_qty = item.get('order_qty', 0)
            unit_cost = float(item.get('unit_cost', 0))
            total = order_qty * unit_cost

            row_data = {
                'upc': item.get('upc', '') or '',
                'description': (item.get('description', '') or '')[:35],
                'on_hand': str(int(item.get('on_hand', 0))),
                'threshold': str(int(item.get('threshold', 0))),
                'suggested_qty': str(int(item.get('suggested_qty', 0))),
                'order_qty': str(int(order_qty)),
                'cases': str(int(item.get('cases', 0))),
                'unit_cost': f"${unit_cost:.2f}",
                'total': f"${total:.2f}"
            }

            table_data.append([row_data[col] for col in selected_columns])

        # Build totals row based on selected columns
        totals_row = []
        for col in selected_columns:
            if col == 'order_qty':
                totals_row.append(str(int(total_units)))
            elif col == 'total':
                totals_row.append(f"${total_cost:.2f}")
            elif col == 'suggested_qty':
                totals_row.append('TOTALS:')
            else:
                totals_row.append('')
        table_data.append(totals_row)

        # Get column widths for selected columns
        col_widths = [column_config[col]['width'] for col in selected_columns]

        table = Table(table_data, colWidths=col_widths)

        # Find description column index for left-alignment
        desc_col_idx = selected_columns.index('description') if 'description' in selected_columns else -1

        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a73e8')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
        ]

        # Left-align description column if present
        if desc_col_idx >= 0:
            table_style.append(('ALIGN', (desc_col_idx, 1), (desc_col_idx, -1), 'LEFT'))

        table.setStyle(TableStyle(table_style))
        elements.append(table)
        doc.build(elements)
        output.seek(0)

        filename = f"inventory-export-{datetime.now().strftime('%Y%m%d')}.pdf"
        return send_file(
            output,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )
    except ImportError:
        return jsonify({'success': False, 'error': 'reportlab not installed'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ============== Tracker History Endpoint ==============

@app.route('/api/tracker/history', methods=['POST'])
def tracker_history():
    try:
        data = request.get_json()
        upcs = data.get('upcs', [])
        months = data.get('months', [])

        if not upcs or not months:
            return jsonify({'success': True, 'data': {}})

        tracker_url = pg.get_setting('tracker_url')
        if not tracker_url:
            return jsonify({'success': False, 'error': 'Tracker URL not configured'}), 400

        tracker_url = tracker_url.rstrip('/')
        result_data = {}

        def fetch_one(upc, month_index, month):
            try:
                url = f"{tracker_url}/api/item-tracker/summary"
                resp = requests.get(url, params={
                    'upc': upc,
                    'from': month['from'],
                    'to': month['to']
                }, timeout=5)
                if resp.status_code == 200:
                    body = resp.json()
                    qty = body.get('quantity_totals', {})
                    return (upc, month_index, {
                        'sale': qty.get('sale', 0) or 0,
                        'purchase': qty.get('purchase', 0) or 0,
                        'beginning_inventory': body.get('beginning_inventory', 0) or 0
                    })
            except Exception:
                pass
            return None

        tasks = []
        for upc in upcs:
            for i, month in enumerate(months):
                tasks.append((upc, i, month))

        with ThreadPoolExecutor(max_workers=20) as executor:
            futures = [executor.submit(fetch_one, upc, i, m) for upc, i, m in tasks]
            for future in futures:
                res = future.result()
                if res:
                    upc, month_index, values = res
                    if upc not in result_data:
                        result_data[upc] = {}
                    result_data[upc][str(month_index)] = values

        return jsonify({'success': True, 'data': result_data})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
